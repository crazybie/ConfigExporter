from src.Utils import *
import csv

import openpyxl


def keyFn(f):
    fName, mode, root = f
    r = ''
    for i in fName:
        if i.isupper() or i == '.':
            return r
        r += i
    return r


def value(v):
    try:
        return int(v)
    except:
        pass
    try:
        return float(v)
    except:
        pass
    v = v.strip()

    if re.search(r'(\w+)-(\w+)-(\w+)', v):
        return re.sub(r'(\w+)-(\w+)-(\w+)', r'\1~\2~\3', v)

    if re.search(r'(\w+)-(\w+)', v):
        return re.sub(r'(\w+)-(\w+)', r'\1~\2', v)
    return v


def isInt(v):
    try:
        int(v)
        return True
    except:
        return False


def isFloat(v):
    try:
        float(v)
        return True
    except:
        return False


def checkType(v):
    if isInt(v):
        return 'int'
    if isFloat(v):
        return 'float'
    if isinstance(v, str) and v != '':
        return 'str'


def appendType(f, tbName, data):
    if f.startswith('comment'): return f

    types = defaultdict(int)

    full = tbName.replace('.csv', '_default.csv')
    defVal = ''
    if os.path.isfile(full):
        csvFile = csv.DictReader(file(full))
        defRow = [i for i in csvFile][0]
        if f in defRow:
            defVal = defRow[f]
            t = checkType(defVal)
            if t:
                types[t] += 1

    for r in data:
        t = checkType(r[f])
        if t:
            types[t] += 1

    t = 'str'
    if types:
        if types['str'] > 0:
            t = 'str'
        elif types['float'] > 0:
            t = 'float'
        elif types['int'] > 0:
            t = 'int'

    if defVal:
        if t == 'str':
            defVal = "='%s'" % defVal
        elif t == 'int':
            defVal = '=%s' % defVal
        elif t == 'float':
            defVal = '=%s' % defVal
    return '%s : %s %s' % (f, t, defVal)


def migrate(csvList):
    for cat, tables in groupBy(csvList, keyFn).iteritems():
        wb = openpyxl.Workbook()
        for t, mode, root in tables:
            tbName = os.path.splitext(t)[0]
            sh = wb.create_sheet(tbName + '@' + mode)
            full = os.path.join(root, t)
            csvFile = csv.DictReader(file(full))

            data = [i for i in csvFile]
            for idx, f, in enumerate(csvFile.fieldnames):
                f = appendType(f, full, data)
                sh.cell(row=1, column=idx + 1, value=f)
            for idx, r in enumerate(data):
                for k, v in r.iteritems():
                    col = csvFile.fieldnames.index(k) + 1
                    sh.cell(row=idx + 2, column=col, value=value(v))

        wb.remove(wb.get_sheet_by_name('Sheet'))
        wb.save('./%s.xlsx' % cat)
        print 'saving', cat


oldCfg = r'E:\EOW\eow-client\GameResource\Config'
oldCfgServer = r'E:\EOW\eow-server\game-server\config\datacsv'

clientCsv = {os.path.basename(f): [os.path.basename(f), 'c', oldCfg] for f in os.listdir(oldCfg) if f.endswith('.csv') and not f.split('.')[0].endswith('_default')}
serverCsv = {os.path.basename(f): [os.path.basename(f), 's', oldCfgServer] for f in os.listdir(oldCfgServer) if f.endswith('.csv') and not f.split('.')[0].endswith('_default')}
for k, v in clientCsv.iteritems():
    if k in serverCsv:
        del serverCsv[k]
        v[1] = 'cs'
migrate(clientCsv.values() + serverCsv.values())
